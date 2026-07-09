#!/usr/bin/env python3
"""
excel-checklist.py — Generate remediation checklist (.xlsx) from findings JSON.

Produces a color-coded Excel workbook with:
  - Findings sheet: ID, Title, Severity (colored), Host, Description, Remediation
  - "Fixed?" column with dropdown (Yes/No/Partial)
  - "Retest Date" column
  - Conditional formatting (green=fixed, red=unfixed)
  - Summary sheet with stats

Usage:
    python excel-checklist.py findings.json --output checklist.xlsx
    python excel-checklist.py findings.json --output checklist.xlsx --update  # Update existing
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

SEVERITY_COLORS = {
    "critical": "CC4141",
    "high": "E85D3F",
    "medium": "F4A442",
    "low": "5B9E5B",
    "info": "4A90D9",
}

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}


def load_findings(path: str) -> list[dict]:
    with open(path) as f:
        findings = json.load(f)
    findings.sort(key=lambda f: SEVERITY_ORDER.get(f.get("severity", "info"), 99))
    return findings


def generate_checklist(findings: list[dict], output_path: str, update: bool = False) -> str:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import CellIsRule, FormulaRule
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return "ERROR: openpyxl not installed. Run: pip install openpyxl"

    if update and Path(output_path).exists():
        wb = load_workbook(output_path)
        ws = wb["Findings"]
        # Append new findings after existing rows
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Findings"
        next_row = 2

    # Header styling (skip if updating existing workbook)
    if not update or not Path(output_path).exists():
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        headers = ["ID", "Severity", "Title", "Host", "Service", "CVE", "CVSS",
                   "Description", "Remediation", "Fixed?", "Retest Date", "Notes"]
        col_widths = [12, 10, 35, 18, 15, 18, 8, 40, 40, 10, 14, 20]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    severity_fills = {sev: PatternFill(start_color=color, end_color=color, fill_type="solid")
                      for sev, color in SEVERITY_COLORS.items()}

    for row_idx, finding in enumerate(findings, next_row):
        data = [
            finding.get("id", ""),
            (finding.get("severity", "") or "").upper(),
            finding.get("title", ""),
            finding.get("host", ""),
            f"{finding.get('service', '')} {finding.get('version', '')}".strip(),
            finding.get("cve", ""),
            finding.get("cvss_score", ""),
            finding.get("description", ""),
            finding.get("remediation", ""),
            finding.get("fixed", ""),
            finding.get("retest_date", ""),
            finding.get("notes", ""),
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_idx == 2:  # Severity column
                sev = finding.get("severity", "info")
                if sev in severity_fills:
                    cell.fill = severity_fills[sev]
                    cell.font = Font(name="Calibri", size=10, bold=True,
                                     color="FFFFFF" if sev in ("critical", "high") else "000000")

    # Data validation — Fixed? dropdown
    dv = DataValidation(type="list", formula1='"Yes,No,Partial,Retest Needed"', allow_blank=True)
    dv.error = "Please select from the dropdown"
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"J2:J{len(findings) + 1}")

    # Conditional formatting for Fixed? column
    if findings:
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        ws.conditional_formatting.add(
            f"A2:L{len(findings) + 1}",
            FormulaRule(formula=[f'$J2="Yes"'], fill=green_fill)
        )
        ws.conditional_formatting.add(
            f"A2:L{len(findings) + 1}",
            FormulaRule(formula=[f'$J2="No"'], fill=red_fill)
        )
        ws.conditional_formatting.add(
            f"A2:L{len(findings) + 1}",
            FormulaRule(formula=[f'OR($J2="Partial",$J2="Retest Needed")'], fill=yellow_fill)
        )

    # Freeze header
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(findings) + 1}"

    # Row heights
    ws.row_dimensions[1].height = 25
    for row_idx in range(2, len(findings) + 2):
        ws.row_dimensions[row_idx].height = 45

    # ── Summary Sheet ──
    ws2 = wb.create_sheet("Summary")

    # Title
    ws2.merge_cells("A1:C1")
    title_cell = ws2.cell(row=1, column=1, value="Remediation Checklist Summary")
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="333333")

    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=2, column=1).font = Font(name="Calibri", size=10, color="797979")

    ws2.cell(row=3, column=1, value=f"Total Findings: {len(findings)}")
    ws2.cell(row=3, column=1).font = Font(name="Calibri", size=12, bold=True)

    # Severity breakdown
    ws2.cell(row=5, column=1, value="Severity").font = Font(bold=True)
    ws2.cell(row=5, column=2, value="Count").font = Font(bold=True)
    ws2.cell(row=5, column=3, value="Fixed").font = Font(bold=True)

    sev_counts = {}
    sev_fixed = {}
    for f in findings:
        sev = f.get("severity", "info")
        sev_counts[sev] = sev_counts.get(sev, 0) + 1
        if f.get("fixed") == "Yes":
            sev_fixed[sev] = sev_fixed.get(sev, 0) + 1

    for row_idx, sev in enumerate(["critical", "high", "medium", "low", "info"], 6):
        if sev not in sev_counts:
            continue
        ws2.cell(row=row_idx, column=1, value=sev.upper())
        ws2.cell(row=row_idx, column=2, value=sev_counts[sev])
        ws2.cell(row=row_idx, column=3, value=sev_fixed.get(sev, 0))

        if sev in severity_fills:
            ws2.cell(row=row_idx, column=1).fill = severity_fills[sev]
            ws2.cell(row=row_idx, column=1).font = Font(
                color="FFFFFF" if sev in ("critical", "high") else "000000", bold=True
            )

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate Excel remediation checklist")
    parser.add_argument("findings", help="JSON file with findings")
    parser.add_argument("--output", required=True, help="Output .xlsx file")
    parser.add_argument("--update", action="store_true", help="Update existing checklist")
    args = parser.parse_args()

    findings = load_findings(args.findings)
    result = generate_checklist(findings, args.output, update=args.update)
    print(f"Checklist: {result}")


if __name__ == "__main__":
    main()
