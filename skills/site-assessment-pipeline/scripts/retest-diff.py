#!/usr/bin/env python3
"""
retest-diff.py — Compare baseline and retest scans, classify findings.

Matches findings by host+port+CVE (or host+service as fallback).
Classifies each finding as: FIXED, STILL_OPEN, NEW, or REGRESSION.
Optionally updates an existing Excel checklist with ticks.

Usage:
    python retest-diff.py baseline.json retest.json                  # Print delta
    python retest-diff.py baseline.json retest.json --update checklist.xlsx  # Update Excel
    python retest-diff.py baseline.json retest.json --json            # JSON output
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path


def load_findings(path: str) -> list[dict]:
    with open(path) as f:
        return json.load(f)


def make_key(finding: dict) -> str:
    """Create a match key from host+port+CVE or fall back to host+service."""
    host = finding.get("host", "")
    port = finding.get("port", "")
    cve = finding.get("cve", "")
    if cve:
        return f"{host}:{port}:{cve}"
    service = finding.get("service", "")
    return f"{host}:{port}:{service}"


def compare(baseline: list[dict], retest: list[dict]) -> dict:
    """Compare two scan result sets. Returns delta report."""
    baseline_keys = {make_key(f): f for f in baseline}
    retest_keys = {make_key(f): f for f in retest}

    fixed = []
    still_open = []
    new = []
    regression = []

    # Findings in baseline but NOT in retest → FIXED
    for key, finding in baseline_keys.items():
        if key not in retest_keys:
            finding["retest_status"] = "FIXED"
            fixed.append(finding)

    # Findings in both baseline and retest → STILL_OPEN
    for key, finding in retest_keys.items():
        if key in baseline_keys:
            baseline_finding = baseline_keys[key]
            finding["retest_status"] = "STILL_OPEN"

            # Check for regression (severity got worse)
            old_sev = _sev_rank(baseline_finding.get("severity", "info"))
            new_sev = _sev_rank(finding.get("severity", "info"))
            if new_sev < old_sev:  # Lower rank = more severe
                finding["retest_status"] = "REGRESSION"
                finding["old_severity"] = baseline_finding.get("severity")
                regression.append(finding)
            else:
                still_open.append(finding)

    # Findings in retest but NOT in baseline → NEW
    for key, finding in retest_keys.items():
        if key not in baseline_keys:
            finding["retest_status"] = "NEW"
            new.append(finding)

    return {
        "scan_date": datetime.now().isoformat(),
        "baseline_total": len(baseline),
        "retest_total": len(retest),
        "fixed": fixed,
        "still_open": still_open,
        "new": new,
        "regression": regression,
        "summary": {
            "fixed_count": len(fixed),
            "still_open_count": len(still_open),
            "new_count": len(new),
            "regression_count": len(regression),
        },
    }


def _sev_rank(severity: str) -> int:
    """Lower = more severe."""
    return {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}.get(severity, 99)


def update_checklist(delta: dict, checklist_path: str) -> str:
    """Update an existing Excel checklist with retest results."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        return "ERROR: openpyxl not installed. Run: pip install openpyxl"

    wb = load_workbook(checklist_path)
    ws = wb["Findings"]

    # Map finding ID to row
    id_to_row = {}
    for row in range(2, ws.max_row + 1):
        fid = ws.cell(row=row, column=1).value
        if fid:
            id_to_row[fid] = row

    # Column indices (must match excel-checklist.py header order)
    COL_FIXED = 10   # "Fixed?" column
    COL_RETEST = 11  # "Retest Date" column
    COL_NOTES = 12   # "Notes" column

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Mark fixed findings
    for finding in delta["fixed"]:
        fid = finding.get("id")
        if fid in id_to_row:
            row = id_to_row[fid]
            ws.cell(row=row, column=COL_FIXED, value="Yes")
            ws.cell(row=row, column=COL_FIXED).fill = green_fill
            ws.cell(row=row, column=COL_RETEST, value=datetime.now().strftime("%Y-%m-%d"))

    # Add new findings
    next_row = ws.max_row + 1
    for finding in delta["new"]:
        data = [
            finding.get("id", f"NEW-{next_row}"),
            (finding.get("severity", "") or "").upper(),
            finding.get("title", ""),
            finding.get("host", ""),
            f"{finding.get('service', '')} {finding.get('version', '')}".strip(),
            finding.get("cve", ""),
            finding.get("cvss_score", ""),
            finding.get("description", ""),
            finding.get("remediation", ""),
            "No",
            "",
            "NEW finding from retest",
        ]
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=next_row, column=col_idx, value=value)
        next_row += 1

    # Update summary
    ws2 = wb["Summary"]
    if delta["baseline_total"] == 0:
        fix_rate = None
        ws2.cell(row=10, column=1, value="Fix Rate: N/A (no baseline findings)")
    else:
        fix_rate = delta["summary"]["fixed_count"] / delta["baseline_total"] * 100
        ws2.cell(row=10, column=1, value=f"Fix Rate: {fix_rate:.0f}%")
    ws2.cell(row=11, column=1, value=f"New Findings: {delta['summary']['new_count']}")
    ws2.cell(row=12, column=1, value=f"Regressions: {delta['summary']['regression_count']}")

    wb.save(checklist_path)
    return checklist_path


def render_text(delta: dict) -> str:
    """Human-readable delta report."""
    s = delta["summary"]
    lines = [
        "Retest Delta Report",
        f"  Baseline: {delta['baseline_total']} findings",
        f"  Retest:   {delta['retest_total']} findings",
        "",
        f"  [OK] FIXED:      {s['fixed_count']}",
        f"  [--] STILL OPEN:  {s['still_open_count']}",
        f"  [+] NEW:          {s['new_count']}",
        f"  [!!] REGRESSION:  {s['regression_count']}",
        "",
    ]

    if delta["fixed"]:
        lines.append("Fixed:")
        for f in delta["fixed"]:
            lines.append(f"  + {f.get('id', '?')}: {f.get('title', '')}")

    if delta["new"]:
        lines.append("\nNew:")
        for f in delta["new"]:
            lines.append(f"  + {f.get('id', '?')}: {f.get('title', '')} ({f.get('severity', '?')})")

    if delta["regression"]:
        lines.append("\nRegressions (severity increased):")
        for f in delta["regression"]:
            lines.append(f"  !! {f.get('id', '?')}: {f.get('title', '')} "
                         f"({f.get('old_severity', '?')} → {f.get('severity', '?')})")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Compare baseline and retest scan findings")
    parser.add_argument("baseline", help="Baseline findings JSON")
    parser.add_argument("retest", help="Retest findings JSON")
    parser.add_argument("--json", action="store_true", help="Output as JSON")
    parser.add_argument("--update", type=str, help="Update Excel checklist with results")
    args = parser.parse_args()

    baseline = load_findings(args.baseline)
    retest = load_findings(args.retest)
    delta = compare(baseline, retest)

    if args.json:
        print(json.dumps(delta, indent=2))
    elif args.update:
        result = update_checklist(delta, args.update)
        print(f"Updated: {result}")
    else:
        print(render_text(delta))

    # Auto-save retest results to scan history DB
    try:
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent / "scripts"))
        from scan_db import ScanDB
        db = ScanDB()
        # Derive engagement ID from filename
        eng_id = Path(args.baseline).stem.replace("findings", "").strip("-_") or "retest"
        for f in delta.get("fixed", []):
            db.save_retest(eng_id, f.get("id", ""), "FIXED")
        for f in delta.get("still_open", []):
            db.save_retest(eng_id, f.get("id", ""), "STILL_OPEN")
        for f in delta.get("new", []):
            db.save_retest(eng_id, f.get("id", "NEW"), "NEW")
        for f in delta.get("regression", []):
            db.save_retest(eng_id, f.get("id", ""), "REGRESSION")
    except Exception:
        pass  # DB save is optional

    # Exit code: non-zero if regressions found
    if delta["summary"]["regression_count"] > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
