"""Tests for excel-checklist.py — remediation checklist generation."""
import importlib.util
import json
import sys
from pathlib import Path

import pytest

# Load excel_checklist module directly by file path (avoid sys.path issues during pytest collection)
SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "skills" / "site-assessment-pipeline" / "scripts"
_excel_path = SCRIPTS_DIR / "excel-checklist.py"
_spec = importlib.util.spec_from_file_location("excel_checklist", str(_excel_path))
_excel_checklist = importlib.util.module_from_spec(_spec)
sys.modules["excel_checklist"] = _excel_checklist
_spec.loader.exec_module(_excel_checklist)

FIXTURES = Path(__file__).resolve().parent / "fixtures"

# ---------------------------------------------------------------------------
# Guard: skip all tests if openpyxl is not installed
# ---------------------------------------------------------------------------
openpyxl_missing = False
try:
    import openpyxl  # noqa: F401
except ImportError:
    openpyxl_missing = True

pytestmark = pytest.mark.skipif(openpyxl_missing, reason="openpyxl not installed")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def load_sample():
    with open(FIXTURES / "sample-findings.json") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------
class TestGenerateChecklist:
    """Test excel-checklist.generate_checklist with sample-findings.json."""

    def test_generates_xlsx_with_findings_and_summary_sheets(self, tmp_path):
        from excel_checklist import generate_checklist

        findings = load_sample()
        out = tmp_path / "checklist.xlsx"
        result = generate_checklist(findings, str(out))

        assert result == str(out)
        assert out.exists()

        wb = openpyxl.load_workbook(out)
        sheet_names = wb.sheetnames
        assert "Findings" in sheet_names
        assert "Summary" in sheet_names

    def test_severity_colors_applied_to_correct_cells(self, tmp_path):
        from excel_checklist import generate_checklist, SEVERITY_COLORS

        findings = load_sample()
        out = tmp_path / "checklist.xlsx"
        generate_checklist(findings, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["Findings"]

        # Build a map of finding ID → severity for cross-check
        expected_sev = {f["id"]: f["severity"] for f in findings}
        expected_hex = {sev: SEVERITY_COLORS[sev] for sev in SEVERITY_COLORS}

        for row_idx, finding in enumerate(findings, 2):
            fid = ws.cell(row=row_idx, column=1).value
            sev_cell = ws.cell(row=row_idx, column=2)  # Severity column
            sev = expected_sev.get(fid)

            assert sev is not None, f"No severity for {fid}"
            actual_rgb = sev_cell.fill.start_color.rgb or ""
            expected = expected_hex[sev]
            # openpyxl 3.1+ adds '00' alpha prefix; handle both
            if actual_rgb.startswith("00") and len(actual_rgb) == 8:
                actual_rgb = actual_rgb[2:]
            assert actual_rgb.upper() == expected.upper(), (
                f"Row {row_idx}: expected {sev} color {expected_hex[sev]}, "
                f"got {sev_cell.fill.start_color.rgb}"
            )

            # Font check
            assert sev_cell.font.bold is True
            if sev in ("critical", "high"):
                assert sev_cell.font.color.rgb == "00FFFFFF", f"Row {row_idx}: expected white font"
            else:
                assert sev_cell.font.color.rgb == "00000000", f"Row {row_idx}: expected black font"

    def test_dropdown_validation_on_fixed_column(self, tmp_path):
        from excel_checklist import generate_checklist

        findings = load_sample()
        out = tmp_path / "checklist.xlsx"
        generate_checklist(findings, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["Findings"]

        validations = ws.data_validations.dataValidation
        assert len(validations) >= 1, "No data validations found"

        # At least one DataValidation covering column J (Fixed?)
        dv_found = False
        for dv in validations:
            if dv.type == "list":
                for rng in dv.sqref.ranges if hasattr(dv.sqref, "ranges") else [dv.sqref]:
                    addr = str(rng)
                    if "J" in addr:
                        dv_found = True
                        # Check the allowed values
                        assert '"Yes,No,Partial,Retest Needed"' in dv.formula1, (
                            f"Expected dropdown values not found in formula1: {dv.formula1}"
                        )
        assert dv_found, "No dropdown validation found on column J"

    def test_conditional_formatting_rules_applied(self, tmp_path):
        from excel_checklist import generate_checklist

        findings = load_sample()
        out = tmp_path / "checklist.xlsx"
        generate_checklist(findings, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["Findings"]

        rules = ws.conditional_formatting
        total_rules = sum(len(cf.rules) for cf in rules)
        assert total_rules >= 3, f"Expected at least 3 conditional formatting rules, got {total_rules}"

        # Extract distinct formula patterns to verify green/red/yellow
        formulas = set()
        for rule in rules:
            for fmt_rule in rule.rules:
                if fmt_rule.formula:
                    formulas.add("".join(fmt_rule.formula))

        # Verify the three expected conditions
        assert any('$J2="Yes"' in f for f in formulas), "Missing 'Yes' (green) rule"
        assert any('$J2="No"' in f for f in formulas), "Missing 'No' (red) rule"
        assert any("Partial" in f or "Retest Needed" in f for f in formulas), (
            "Missing 'Partial/Retest Needed' (yellow) rule"
        )

    def test_header_row_has_correct_columns(self, tmp_path):
        from excel_checklist import generate_checklist

        findings = load_sample()
        out = tmp_path / "checklist.xlsx"
        generate_checklist(findings, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["Findings"]

        expected_headers = [
            "ID", "Severity", "Title", "Host", "Service", "CVE", "CVSS",
            "Description", "Remediation", "Fixed?", "Retest Date", "Notes",
        ]
        for col_idx, expected in enumerate(expected_headers, 1):
            actual = ws.cell(row=1, column=col_idx).value
            assert actual == expected, f"Column {col_idx}: expected {expected!r}, got {actual!r}"

    def test_data_rows_count_matches_findings(self, tmp_path):
        from excel_checklist import generate_checklist

        findings = load_sample()
        out = tmp_path / "checklist.xlsx"
        generate_checklist(findings, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["Findings"]

        # +1 for header row
        assert ws.max_row == len(findings) + 1, (
            f"Expected {len(findings) + 1} rows, got {ws.max_row}"
        )

    def test_summary_sheet_has_total_and_severity_breakdown(self, tmp_path):
        from excel_checklist import generate_checklist

        findings = load_sample()
        out = tmp_path / "checklist.xlsx"
        generate_checklist(findings, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        # Cell A3 should contain total findings
        total_cell = ws.cell(row=3, column=1).value
        assert total_cell is not None
        assert f"Total Findings: {len(findings)}" in str(total_cell)

        # Severity breakdown header row
        assert ws.cell(row=5, column=1).value == "Severity"
        assert ws.cell(row=5, column=2).value == "Count"
        assert ws.cell(row=5, column=3).value == "Fixed"

    def test_freeze_panes_and_auto_filter_set(self, tmp_path):
        from excel_checklist import generate_checklist

        findings = load_sample()
        out = tmp_path / "checklist.xlsx"
        generate_checklist(findings, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb["Findings"]

        # Freeze panes
        assert ws.freeze_panes == "A2", f"Expected freeze_panes='A2', got {ws.freeze_panes!r}"

        # Auto filter
        assert ws.auto_filter.ref is not None, "Auto filter not set"
